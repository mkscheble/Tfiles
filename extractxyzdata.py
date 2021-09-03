from PIL import Image
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import Button
from matplotlib import pyplot as plt


"""Extracts Offset Data from Tiff Images, Monolayer"""

def getdirectory():
    directory = filedialog.askdirectory()
    print(directory)
    image_directory = r"C:\Users\MarkScheble.Jr\Desktop\imagesnew"
    array = main_func(directory, image_directory, False, imagetype='.png')
    write_data(array, 'temp.xlsx', directory)


def loop(directory, filename):
    # get directory path
    path_name = os.path.join(directory, filename)
    print(directory)
    print(filename)
    try:
        im = Image.open(path_name)
    except:
        return [0,0,0]
    # access description of image
    description = im.tag
    description = description.__str__()

    # extract XOffset
    xpartitioned = description.partition('"XOffset":"')[2]
    xpartitioned = xpartitioned.partition('","YOffset"')[0]

    # extract YOffset
    ypartitioned = description.partition('"YOffset":"')[2]
    ypartitioned = ypartitioned.partition('","ZOffset"')[0]

    # extract ZOffset
    zpartitioned = description.partition('"ZOffset":"')[2]
    zpartitioned = zpartitioned.partition('","FOV"')[0]

    # convert to int
    try:
        XOffset = int(xpartitioned)/1e6
        YOffset = int(ypartitioned)/1e6
        ZOffset = int(zpartitioned)/1e6
    except ValueError:
        XOffset = 0
        YOffset = 0
        ZOffset = 0
    return [XOffset, YOffset, ZOffset]

#saves images as different file format and extract offsets
def main_func(directory, saved_directory, save_images, imagetype='.png'):
    # autofocus = np.array([])
    array = []
    index = -1
    for filename in os.listdir(directory):
        #only files with tiff
        if filename.endswith(".tiff"):
            image = filename.partition("Image[")[2]
            image = image.partition("]")[0]
            image = int(image)

            if index < image:
                index = image
                Offsets = loop(directory, filename)

                # add to big array
                array += [Offsets]
                print("Image:", index, "Offsets (mm) - x:", Offsets[0], "y:", Offsets[1], "z:", Offsets[2])

                filename = os.path.join(directory, filename)
                if save_images:
                    im = Image.open(filename)
                    description = filename.partition("1.0.14/")[2]
                    # description = description.partition("1.0.14/")[2]
                    description = description.partition(".tiff")[0]
                    description = description + imagetype
                    path_new = os.path.join(saved_directory, description)
                    print(path_new)
                    im.save(path_new)

            else:
                continue
        # uncomment for initial autofocus data
        #     filename = os.path.join(directory, filename)
        #     print(filename)
        #     for file in os.listdir(filename):
        #         if file.endswith(".tiff"):
        #             print(file)
        #             Offset = loop(filename, file)
        #             print(Offset)
        #             autofocus = np.append(autofocus, Offset)

        else:
            continue

    return array

def write_data(array, filename, directory):
    #load excel file, need to have one named ZOffsetData
    try:
        wb = load_workbook(filename= filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    col = ws.max_column
    row = 1
    if col == 1:
        ws.cell(row,col, 'Image Number')
        for x in np.arange(130):
            row += 1
            ws.cell(row, col, x+1)
        col = 2
        row = 1
    else:
        col += 1
    #write to excel file
    ws.cell(row, col, directory)

    # Uncomment to add shorten directory label to excel sheet
    # col += 1
    # path = directory.partition("/1.0.")[2]
    # path = path.partition("/")[2]
    # ws.cell(row, col, path)

    ws.cell(row+1,col, 'xoffsets')
    ws.cell(row+1, col+1, 'yoffsets')
    ws.cell(row+1, col+2, 'zoffsets')
    row += 2
    for item in array:
        x,y,z = item
        ws.cell(row, col, x)
        ws.cell(row, col+1,y)
        ws.cell(row,col+2, z)
        row += 1

    while True:
        try:
            wb.save(filename)
            print("saved")
            break
        except:
            print("Close Window")

#Create GUI
root = tk.Tk()
root.title('Offset Extractor')
root.geometry("400x400")
root.configure(bg='green')
button = Button(root, text="Choose a Folder", command=getdirectory, width=12, height=4, font='Times 14 bold', bg='white', fg='#56b5fd')
button.place(x=142, y=142)

root.mainloop()

