from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
from scipy.linalg import null_space


#loads the workbook in and accesses the right sheet
def load_wb(filename, sheetname):
    try:
        wb = load_workbook(filename)
        ws = wb['Sheet']
        print("Worksheet found")
        return ws
    except:
        print("No Workbook Found!")
        return None

#loads and returns the data desired
def accessData(ws):
    row = 3
    column_num = 2
    xyz = []
    array = []
    while True:
        if row > 200:
            column_num += 3
            row = 3
            if len(array) > 0:
                xyz += [array]
            array = []
            if column_num > 200:
                break
        cellx = str(get_column_letter(column_num)) + str(row)
        x = ws[cellx].value
        celly = str(get_column_letter(column_num + 1)) + str(row)
        y = ws[celly].value
        cellz = str(get_column_letter(column_num + 2)) + str(row)
        z = ws[cellz].value
        if x != None and y != None and z != None:
            array += [[ws[cellx].value, ws[celly].value, ws[cellz].value]]
        row += 1
    return xyz

def analyzedata(data):
    arr = []
    for fov in data:
        firstpt = fov[0]
        secondpt = fov[20]
        thirdpt = fov[len(fov) - 1]
        matrix = np.array([[firstpt[0], firstpt[1], firstpt[2], 1], [secondpt[0], secondpt[1], secondpt[2], 1], [thirdpt[0], thirdpt[1], thirdpt[2], 1]])
        ns = null_space(matrix)
        ns = ns / ns[3]
        arr = arr + [ns]
        print(arr)
    return arr

def plotter(planes):
    pass

def find_channel_change(array):
    max_jump = 0
    max_arg = np.inf
    for i in np.arange(len(array)-1):
        temp = np.abs(array[i] - array[i+1])
        if temp > max_jump:
            max_arg = i+1
            max_jump = temp
    if max_jump != 0 and max_arg != np.inf and max_jump > 0.005:
        return max_arg
    else:
        return False


def main():
    path = r"C:\Users\MarkScheble.Jr\Desktop\ZOffsetTool\XYXOffsetDataShortGlass.xlsx"
    sheetname = 'Sheet'
    ws = load_wb(path, sheetname)
    data = accessData(ws)
    # planes = analyzedata(data)
    # plotter(planes)

if __name__ == "__main__":
    main()